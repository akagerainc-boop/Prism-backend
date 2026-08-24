"""Excel (.xlsx) export of every table in the document, via openpyxl.

One sheet per detected table, plus an index sheet that says where each table
came from -- a workbook of anonymous "Sheet1..Sheet9" tabs is useless when the
source was a 40-page scan.
"""

from __future__ import annotations

import io
import re

from ..document_model import table_rows
from ..logging_config import get_logger
from ..schemas import StructuredDocument

log = get_logger(__name__)

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


class ExportUnavailable(RuntimeError):
    """A required export library is not installed."""


def _import_openpyxl():
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
        from openpyxl.styles import Alignment, Font  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore

        return Workbook, Font, Alignment, get_column_letter
    except Exception as exc:
        raise ExportUnavailable(
            "openpyxl is not installed. Run `pip install openpyxl`."
        ) from exc


def _safe_sheet_title(title: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("-", title).strip() or "Table"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        tail = f"_{suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _coerce(value: str):
    """Write numbers as numbers so Excel can actually compute with them."""
    text = (value or "").strip()
    if not text:
        return None
    candidate = text.replace(",", "").replace("%", "").strip()
    try:
        if re.fullmatch(r"[-+]?\d+", candidate):
            return int(candidate)
        if re.fullmatch(r"[-+]?\d*\.\d+([eE][-+]?\d+)?", candidate):
            return float(candidate)
    except ValueError:  # pragma: no cover
        pass
    return text


def export_xlsx(document: StructuredDocument, **_: object) -> bytes:
    """Render every detected table into an .xlsx workbook."""
    Workbook, Font, Alignment, get_column_letter = _import_openpyxl()

    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "Index"
    index_sheet.append(["Sheet", "Page", "Rows", "Columns", "Caption"])
    for cell in index_sheet[1]:
        cell.font = Font(bold=True)

    used_titles: set[str] = {"index"}
    table_count = 0

    for page in document.pages:
        elements = sorted(page.elements, key=lambda e: e.readingOrder)
        for element in elements:
            if element.type != "table" or element.table is None:
                continue

            rows = table_rows(element.table)
            if not rows:
                continue

            table_count += 1

            # A caption attached to this table names the sheet.
            caption = next(
                (
                    (c.text or "").strip()
                    for c in elements
                    if c.type == "caption" and c.parentId == element.id and c.text
                ),
                "",
            )
            title = _safe_sheet_title(
                caption or f"P{page.pageNumber}_Table{table_count}", used_titles
            )
            sheet = workbook.create_sheet(title=title)

            for row in rows:
                sheet.append([_coerce(cell) for cell in row])

            header_flagged = any(c.isHeader and c.row == 0 for c in element.table.cells)
            if header_flagged and sheet.max_row >= 1:
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                sheet.freeze_panes = "A2"

            # Re-apply merged cells.
            for cell_spec in element.table.cells:
                if cell_spec.rowSpan <= 1 and cell_spec.colSpan <= 1:
                    continue
                try:
                    sheet.merge_cells(
                        start_row=cell_spec.row + 1,
                        start_column=cell_spec.col + 1,
                        end_row=cell_spec.row + cell_spec.rowSpan,
                        end_column=cell_spec.col + cell_spec.colSpan,
                    )
                except Exception as exc:  # pragma: no cover
                    log.debug("Skipped an xlsx merge: %s", exc)

            # Width heuristic so the sheet is readable without manual resizing.
            for column_index in range(1, (sheet.max_column or 0) + 1):
                longest = 0
                for row in sheet.iter_rows(
                    min_col=column_index, max_col=column_index, values_only=True
                ):
                    value = row[0]
                    if value is not None:
                        longest = max(longest, len(str(value)))
                sheet.column_dimensions[get_column_letter(column_index)].width = min(
                    max(10, longest + 2), 60
                )

            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            index_sheet.append(
                [
                    title,
                    page.pageNumber,
                    len(rows),
                    max(len(r) for r in rows),
                    caption,
                ]
            )

    if table_count == 0:
        index_sheet.append(["No tables were detected in this document.", "", "", "", ""])

    for column_index, width in enumerate((28, 8, 8, 10, 50), start=1):
        index_sheet.column_dimensions[get_column_letter(column_index)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
